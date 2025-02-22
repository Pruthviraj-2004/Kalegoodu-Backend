from django.shortcuts import render

def home(request):
    return render(request, 'kalegooduapp/home.html')

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import BannerImage, Category, PageContent, PageImage, SaleType, Product, ProductImage, CategoryImage, Comment, Customer, Order, OrderItem, Workshop, WorkshopImage, WorkshopVideo
from .serializers import *
from django.shortcuts import get_object_or_404
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework import status
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from .utils import send_whatsapp_message
from django.contrib.auth import authenticate, login, logout
from cloudinary.uploader import destroy
from rest_framework.pagination import PageNumberPagination

from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string

from django.views import View
from openpyxl import Workbook
from datetime import datetime

from django.db.models import Q, Case, When, F, Value, IntegerField, Count

import razorpay
from django.conf import settings

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 4
    page_size_query_param = 'page_size'
    max_page_size = 50

class AdminStandardResultsSetPagination(PageNumberPagination):
    page_size = 4
    page_size_query_param = 'page_size'
    max_page_size = 50

# @method_decorator(csrf_exempt, name='dispatch')
# class SaleTypeView(APIView):
#     def get(self, request):
#         sale_types = SaleType.objects.all()
#         serializer = SaleTypeSerializer(sale_types, many=True)
#         return Response({'sale_types': serializer.data})
@method_decorator(csrf_exempt, name='dispatch')
class PageImageUpdateView(APIView):
    def put(self, request, pageimage_id):
        try:
            # Get the PageImage instance
            page_image = PageImage.objects.get(pk=pageimage_id)

            # Check if a new image is provided
            new_image = request.data.get('image')

            if new_image:
                # Delete the old image from Cloudinary (if exists)
                if page_image.image:
                    public_id = page_image.image.public_id
                    destroy(public_id)

                # Update with new image
                page_image.image = new_image
                page_image.save()

            return Response({"message": "Image updated successfully"}, status=status.HTTP_200_OK)

        except PageImage.DoesNotExist:
            return Response({"error": "PageImage not found"}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class SaleTypeView(APIView):
    def get(self, request):
        search_query = request.query_params.get('search', '')
        sort_by = request.query_params.get('sort_by', 'name')
        sort_order = request.query_params.get('sort_order', 'asc')

        filters = Q()
        if search_query:
            filters &= Q(name__icontains=search_query)

        sale_types = SaleType.objects.filter(filters)

        sort_fields = []
        if sort_by == 'name':
            sort_fields.append('name' if sort_order == 'asc' else '-name')
        elif sort_by == 'visible':
            if sort_order == 'true':
                sort_fields.append('-visible')
            else:
                sort_fields.append('visible')
            sort_fields.append('name')

        if not sort_fields:
            sort_fields.append('name')

        sale_types = sale_types.order_by(*sort_fields)

        serializer = SaleTypeSerializer(sale_types, many=True)

        return Response({'sale_types': serializer.data})

    def post(self, request):
        serializer = SaleTypeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'sale_type': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryView(APIView):
    def get(self, request):
        # Extract query parameters for filtering and sorting
        search_query = request.query_params.get('search', '')
        sort_by = request.query_params.get('sort_by', 'created_at')
        sort_order = request.query_params.get('sort_order', 'asc')
        visible_filter = request.query_params.get('visible', None)  # Filter for visible field

        # Build the query for filtering
        filters = Q()
        if search_query:
            filters &= Q(name__icontains=search_query)
        if visible_filter is not None:  # Filter based on visible field if parameter is provided
            filters &= Q(visible=visible_filter.lower() in ['true', '1'])

        # Query the categories with the applied filters
        categories = Category.objects.filter(filters)

        # Determine sorting field and order
        sort_fields = []
        if sort_by == 'name':
            sort_fields.append('name' if sort_order == 'asc' else '-name')
        elif sort_by == 'created_at':
            sort_fields.append('created_at' if sort_order == 'asc' else '-created_at')
        elif sort_by == 'visible':  # Sorting by visible field
            if sort_order == 'true':
                sort_fields.append('-visible')  # False (hidden) first, then True (visible)
            else:
                sort_fields.append('visible')  # True (visible) first, then False (hidden)
            sort_fields.append('name')  # Secondary sort by name
        else:
            sort_fields.append('created_at')  # Default sort by created_at

        categories = categories.order_by(*sort_fields)

        # Serialize the categories
        serializer = CategorySerializer(categories, many=True)

        # Return the response
        return Response({'categories': serializer.data})

    def post(self, request):
        serializer = CategorySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'category': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class VisibleCategoryView(APIView):
    def get(self, request):
        filters = Q(visible=True,home_page=True)

        categories = Category.objects.filter(filters)
        serializer = CategorySerializer(categories, many=True)

        return Response({'categories': serializer.data})

class VisibleCategoryHeaderView(APIView):
    def get(self, request):
        filters = Q(visible=True,header=True)

        categories = Category.objects.filter(filters)
        serializer = SimpleCategorySerializer(categories, many=True)

        return Response({'categories': serializer.data})


# @method_decorator(csrf_exempt, name='dispatch')
# class ProductView(APIView):
#     def get(self, request):
#         products = Product.objects.all()
#         serializer = ProductSerializer(products, many=True)
#         return Response({'products': serializer.data})

#     def post(self, request):
#         serializer = ProductSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response({'product': serializer.data}, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class ProductView(APIView):
    pagination_class = StandardResultsSetPagination

    def get(self, request):
        search_query = request.query_params.get('search', '')
        sort_by = request.query_params.get('sort_by', 'effective_price')
        sort_order = request.query_params.get('sort_order', 'asc')
        visible_filter = request.query_params.get('visible', None)

        filters = Q()
        if search_query:
            filters &= Q(name__icontains=search_query)

        # Apply the visible filter if provided
        if visible_filter is not None:
            filters &= Q(visible=visible_filter.lower() in ['true', '1'])

        try:
            products = Product.objects.filter(filters).annotate(
                effective_price=Case(
                    When(discounted_price=0, then=F('price')),
                    default=F('discounted_price'),
                    output_field=IntegerField()
                )
            )

            # Sorting logic
            sort_fields = []
            if sort_by == 'name':
                sort_fields.append('name' if sort_order == 'asc' else '-name')
            elif sort_by == 'created_at':
                sort_fields.append('created_at' if sort_order == 'asc' else '-created_at')
            elif sort_by == 'visible':
                if sort_order == 'true':
                    sort_fields.append('-visible')  # False (hidden) first, then True (visible)
                else:
                    sort_fields.append('visible')  # True (visible) first, then False (hidden)
                sort_fields.append('name')  # Secondary sort by name
            else:
                sort_fields.append('effective_price' if sort_order == 'asc' else '-effective_price')

            # Apply the sorting
            products = products.order_by(*sort_fields)

            paginator = self.pagination_class()
            paginated_products = paginator.paginate_queryset(products, request)

            serializer = ProductSerializer(paginated_products, many=True)
            return paginator.get_paginated_response({'products': serializer.data})

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def post(self, request):
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'product': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class ListProductView(APIView):
#     def get(self, request):
#         products = Product.objects.all()

#         serializer = NewProductSerializer(products, many=True)

#         return Response({'products': serializer.data}, status=status.HTTP_200_OK)

class ListProductView(APIView):
    def get(self, request, *args, **kwargs):
        search_query = request.query_params.get('search', '')
        min_price = request.query_params.get('min_price')
        max_price = request.query_params.get('max_price')
        sort_by = request.query_params.get('sort_by', 'effective_price')
        sort_order = request.query_params.get('sort_order', 'asc')

        filters = Q(visible=True)
        if search_query:
            filters &= Q(name__icontains=search_query)
        if min_price:
            filters &= Q(discounted_price__gte=min_price)
        if max_price:
            filters &= Q(discounted_price__lte=max_price)

        try:
            products = Product.objects.filter(filters).annotate(
                effective_price=Case(
                    When(discounted_price=0, then=F('price')),
                    default=F('discounted_price'),
                    output_field=IntegerField()
                )
            )

            if sort_by == 'name':
                sort_field = 'name' if sort_order == 'asc' else '-name'
            else:
                sort_field = 'effective_price' if sort_order == 'asc' else '-effective_price'

            products = products.order_by(sort_field)

            paginator = StandardResultsSetPagination()
            paginated_products = paginator.paginate_queryset(products, request)

            serializer = ProductSerializer(paginated_products, many=True)
            return paginator.get_paginated_response(serializer.data)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

@method_decorator(csrf_exempt, name='dispatch')
class CategoryImageView(APIView):
    def get(self, request):
        category_images = CategoryImage.objects.all()
        serializer = CategoryImageSerializer(category_images, many=True)
        return Response({'category_images': serializer.data})

    def post(self, request):
        serializer = CategoryImageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'category_image': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class ProductImageView(APIView):
    def get(self, request):
        product_images = ProductImage.objects.all()
        serializer = ProductImageSerializer(product_images, many=True)
        return Response({'product_images': serializer.data})

    def post(self, request):
        serializer = ProductImageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'product_image': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class AllCommentView(APIView):
    def get(self, request):
        # Get the search query parameter
        search_query = request.query_params.get('search', '')

        # Filter comments based on the search query
        if search_query:
            comments = Comment.objects.filter(
                Q(user_name__icontains=search_query)    # Example: If you have an 'author' field
            )
        else:
            comments = Comment.objects.all()

        # Serialize the filtered comments
        serializer = CommentSerializer(comments, many=True)
        return Response({'comments': serializer.data})

    def post(self, request):
        serializer = CommentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'comment': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class CommentView(APIView):
    def get(self, request):
        comments = Comment.objects.filter(display=True).order_by('-updated_at')
        serializer = CommentSerializer(comments, many=True)
        return Response({'comments': serializer.data})

@method_decorator(csrf_exempt, name='dispatch')
class CategoryDetailAPIView(APIView):
    def get(self, request, category_id):
        category = get_object_or_404(Category, pk=category_id)
        serializer = CategorySerializer(category)
        return Response({'category': serializer.data})

    def post(self, request):
        serializer = CategorySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'category': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class SaleTypeDetailAPIView(APIView):
    def get(self, request, sale_type_id):
        sale_type = get_object_or_404(SaleType, pk=sale_type_id)
        serializer = SaleTypeSerializer(sale_type)
        return Response({'sale_type': serializer.data})

    def post(self, request):
        serializer = SaleTypeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'sale_type': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class ProductDetailAPIView(APIView):
    def get(self, request, product_id):
        product = get_object_or_404(Product, pk=product_id)
        serializer = ProductSerializer(product)
        return Response({'product': serializer.data})

    def post(self, request):
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'product': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class ProductImageDetailAPIView(APIView):
    def get(self, request, product_image_id):
        product_image = get_object_or_404(ProductImage, pk=product_image_id)
        serializer = ProductImageSerializer(product_image)
        return Response({'product_image': serializer.data})

    def post(self, request):
        serializer = ProductImageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'product_image': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryImageDetailAPIView(APIView):
    def get(self, request, category_image_id):
        category_image = get_object_or_404(CategoryImage, pk=category_image_id)
        serializer = CategoryImageSerializer(category_image)
        return Response({'category_image': serializer.data})

    def post(self, request):
        serializer = CategoryImageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'category_image': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CommentDetailAPIView(APIView):
    def get(self, request, comment_id):
        comment = get_object_or_404(Comment, pk=comment_id)
        serializer = CommentSerializer(comment)
        return Response({'comment': serializer.data})

    def post(self, request):
        serializer = CommentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'comment': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CustomerView(APIView):
    def get(self, request):
        customers = Customer.objects.all()
        serializer = CustomerSerializer(customers, many=True)
        return Response({'customers': serializer.data})

    def post(self, request):
        serializer = CustomerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'customer': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# class ListOrderView(APIView):
#     def get(self, request):
#         orders = Order.objects.all()

#         paginator = AdminStandardResultsSetPagination()
#         paginated_orders = paginator.paginate_queryset(orders, request)

#         serializer = OrderSerializer(paginated_orders, many=True)

#         return paginator.get_paginated_response(serializer.data)

class ListOrderView(APIView):
    def get(self, request):
        order_completed = request.query_params.get('order_completed')
        order_id = request.query_params.get('order_id')
        customer_name = request.query_params.get('customer_name')

        filters = Q()

        if order_completed is not None:
            try:
                order_completed = bool(int(order_completed))
                filters &= Q(order_completed=order_completed)
            except ValueError:
                return Response(
                    {"error": "'order_completed' must be either 0 or 1 (True/False)"},
                    status=status.HTTP_400_BAD_REQUEST
                )

        if order_id:
            try:
                filters &= Q(order_id=order_id)
            except ValueError:
                return Response(
                    {"error": "'order_id' must be an integer."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        if customer_name:
            filters &= Q(customer__name__icontains=customer_name)

        orders = Order.objects.filter(filters).order_by('-order_id')

        paginator = AdminStandardResultsSetPagination()
        paginated_orders = paginator.paginate_queryset(orders, request)

        serializer = OrderSerializer(paginated_orders, many=True)

        return paginator.get_paginated_response(serializer.data)

@method_decorator(csrf_exempt, name='dispatch')
class OrderView(APIView):
    def get(self, request):
        orders = Order.objects.all()
        serializer = OrderSerializer(orders, many=True)
        return Response({'orders': serializer.data})

    def post(self, request):
        serializer = OrderSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'order': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class OrderItemView(APIView):
    def get(self, request):
        order_items = OrderItem.objects.all()
        serializer = OrderItemSerializer(order_items, many=True)
        return Response({'order_items': serializer.data})

    def post(self, request):
        serializer = OrderItemSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'order_item': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class OrderDetailAPIView(APIView):
    def get(self, request, order_id):
        order = get_object_or_404(Order, pk=order_id)
        serializer = OrderSerializer(order)
        return Response({'order': serializer.data})

    def post(self, request):
        serializer = OrderSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'order': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CustomerDetailAPIView(APIView):
    def get(self, request, customer_id):
        customer = get_object_or_404(Customer, pk=customer_id)
        serializer = CustomerSerializer(customer)
        return Response({'customer': serializer.data})

    def post(self, request):
        serializer = CustomerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'customer': serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class OrderDetailWithCustomerAPIView(APIView):
    def get(self, request, order_id):
        order = get_object_or_404(Order, pk=order_id)

        order_serializer = OrderSerializer(order)
        customer_serializer = CustomerSerializer(order.customer)

        return Response({'customer': customer_serializer.data,'order': order_serializer.data})

class PageContentListView(APIView):
    def get(self, request):
        page_contents = PageContent.objects.all()
        serializer = PageContentSerializer(page_contents, many=True)
        return Response({'page_contents': serializer.data}, status=status.HTTP_200_OK)

class PageContentDetailView(APIView):
    def get(self, request, pagecontent_id):
        page_content = get_object_or_404(PageContent, pk=pagecontent_id)
        serializer = PageContentSerializer(page_content)
        return Response({'page_content': serializer.data}, status=status.HTTP_200_OK)

class PageImageListView(APIView):
    def get(self, request):
        page_images = PageImage.objects.all()
        serializer = PageImageSerializer(page_images, many=True)
        return Response({'page_images': serializer.data}, status=status.HTTP_200_OK)

class PageImageDetailView(APIView):
    def get(self, request, pageimage_id):
        page_image = get_object_or_404(PageImage, pk=pageimage_id)
        serializer = PageImageSerializer(page_image)
        return Response({'page_image': serializer.data}, status=status.HTTP_200_OK)


class WorkshopListView(APIView):
    def get(self, request):
        try:
            # Retrieve the search query parameter from the request
            search_term = request.query_params.get('search', None)

            # Start with all workshops
            workshops = Workshop.objects.all()

            # Apply search filter if search term is provided
            if search_term:
                workshops = workshops.filter(
                    Q(name__icontains=search_term) |  # Search in name
                    Q(place__icontains=search_term)   # Search in place
                )

            # Serialize the filtered or unfiltered workshops
            serializer = WorkshopSerializer(workshops, many=True)

            # Return the serialized data
            return Response({'workshops': serializer.data}, status=status.HTTP_200_OK)

        except Exception as e:
            # Generic error handling
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



class WorkshopDetailView(APIView):
    def get(self, request, workshop_id):
        workshop = get_object_or_404(Workshop, pk=workshop_id)
        serializer = WorkshopSerializer(workshop)
        return Response({'workshop': serializer.data}, status=status.HTTP_200_OK)

class WorkshopImageListView(APIView):
    def get(self, request):
        workshop_images = WorkshopImage.objects.all()
        serializer = WorkshopImageSerializer(workshop_images, many=True)
        return Response({'workshop_images': serializer.data}, status=status.HTTP_200_OK)

class WorkshopImageDetailView(APIView):
    def get(self, request, workshop_image_id):
        workshop_image = get_object_or_404(WorkshopImage, pk=workshop_image_id)
        serializer = WorkshopImageSerializer(workshop_image)
        return Response({'workshop_image': serializer.data}, status=status.HTTP_200_OK)

class WorkshopVideoListView(APIView):
    def get(self, request):
        workshop_videos = WorkshopVideo.objects.all()
        serializer = WorkshopVideoSerializer(workshop_videos, many=True)
        return Response({'workshop_videos': serializer.data}, status=status.HTTP_200_OK)

class WorkshopVideoDetailView(APIView):
    def get(self, request, workshop_video_id):
        workshop_video = get_object_or_404(WorkshopVideo, pk=workshop_video_id)
        serializer = WorkshopVideoSerializer(workshop_video)
        return Response({'workshop_video': serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryCreateView(APIView):
    @transaction.atomic
    def post(self, request):
        category_data = {
            'name': request.data.get('name'),
            'description': request.data.get('description'),
            'visible': request.data.get('visible'),
            'header': request.data.get('header'),
            'home_page': request.data.get('home_page')
        }
        category_serializer = CategorySerializer(data=category_data)

        if category_serializer.is_valid():
            category = category_serializer.save()

            images = request.FILES.getlist('images')
            for image in images:
                category_image_data = {
                    'category': category.category_id,
                    'image': image,
                    'alt_text': request.data.get('alt_text', ''),
                    'visible': request.data.get('visible')
                }
                category_image_serializer = CategoryImageSerializer(data=category_image_data)
                if category_image_serializer.is_valid():
                    category_image_serializer.save()
                else:
                    transaction.set_rollback(True)
                    return Response(category_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'category': category_serializer.data}, status=status.HTTP_201_CREATED)

        return Response(category_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class ProductCreateView(APIView):
    @transaction.atomic
    def post(self, request):
        product_data = {
            'name': request.data.get('name'),
            'price': request.data.get('price'),
            'discounted_price': request.data.get('discounted_price', '0'),
            'short_description': request.data.get('short_description'),
            'quantity': request.data.get('quantity'),
            'video_link': request.data.get('video_link'),
            'visible': request.data.get('visible')
        }

        categories_data = request.data.get('categories', '[]')
        sale_types_data = request.data.get('sale_types', '[]')

        try:
            if isinstance(categories_data, str):
                categories_data = json.loads(categories_data)
            if isinstance(sale_types_data, str):
                sale_types_data = json.loads(sale_types_data)
        except json.JSONDecodeError:
            return Response({'error': 'Invalid data format for categories or sale types.'}, status=status.HTTP_400_BAD_REQUEST)

        images_data = request.FILES.getlist('images')

        product_serializer = ProductSerializer(data=product_data)
        if product_serializer.is_valid():
            product = product_serializer.save()

            if categories_data:
                for category_id in categories_data:
                    try:
                        category = Category.objects.get(pk=int(category_id))
                        product.categories.add(category)
                    except (Category.DoesNotExist, ValueError):
                        transaction.set_rollback(True)
                        return Response(
                            {'error': f'Category with ID {category_id} does not exist or invalid.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

            if sale_types_data:
                for sale_type_id in sale_types_data:
                    try:
                        sale_type = SaleType.objects.get(pk=int(sale_type_id))
                        product.sale_types.add(sale_type)
                    except (SaleType.DoesNotExist, ValueError):
                        transaction.set_rollback(True)
                        return Response(
                            {'error': f'SaleType with ID {sale_type_id} does not exist or invalid.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

            for image in images_data:
                product_image_data = {
                    'product': product.product_id,
                    'image': image,
                    'alt_text': request.data.get('alt_text', ''),
                    'visible': request.data.get('visible')
                }
                product_image_serializer = ProductImageSerializer(data=product_image_data)
                if product_image_serializer.is_valid():
                    product_image_serializer.save()
                else:
                    product.delete()
                    transaction.set_rollback(True)
                    return Response(product_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'product': product_serializer.data}, status=status.HTTP_201_CREATED)

        return Response(product_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class BannerImageView(APIView):
    def get(self, request):
        banner_images = BannerImage.objects.all()
        serializer = BannerImageSerializer(banner_images, many=True)
        return Response({'banner_images': serializer.data})

    def post(self, request):
        image = request.FILES.get('image')
        title = request.data.get('title')
        visible = request.data.get('visible')

        if not image:
            return Response({'error': 'No image file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        banner_image_data = {
            'title': title,
            'image': image,
            'visible': visible
        }

        serializer = BannerImageSerializer(data=banner_image_data)
        if serializer.is_valid():
            serializer.save()
            return Response({'banner_image': serializer.data}, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class PageContentCreateView(APIView):
    @transaction.atomic
    def post(self, request):
        page_content_data = {
            'page_name': request.data.get('page_name'),
            'content': request.data.get('content'),
            'visible': request.data.get('visible')
        }
        page_content_serializer = PageContentSerializer(data=page_content_data)

        if page_content_serializer.is_valid():
            page_content = page_content_serializer.save()

            page_images = request.FILES.getlist('page_image')
            for image in page_images:
                page_image_data = {
                    'page': page_content.pagecontent_id,
                    'image': image,
                    'visible': request.data.get('visible')
                }
                page_image_serializer = PageImageSerializer(data=page_image_data)
                if page_image_serializer.is_valid():
                    page_image_serializer.save()
                else:
                    transaction.set_rollback(True)
                    return Response(page_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'page_content': page_content_serializer.data}, status=status.HTTP_201_CREATED)

        return Response(page_content_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class CategoriesByProductView(APIView):
    def get(self, request, product_id):
        try:
            product = Product.objects.get(product_id=product_id)
            categories = product.categories.all()
            serializer = CategorySerializer(categories, many=True)
            return Response({'categories': serializer.data}, status=status.HTTP_200_OK)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

class SubCategoryListByCategoryView(APIView):
    def get(self, request, category_id):
        try:
            category = Category.objects.get(category_id=category_id)
            subcategories = SubCategory.objects.filter(category__category_id=category_id)
            subcategory_serializer = SubCategorySerializer(subcategories, many=True)

            return Response({'subcategories': subcategory_serializer.data}, status=status.HTTP_200_OK)
        except Category.DoesNotExist:
            return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

class ProductsByCategoryView(APIView):
    def get(self, request, category_id):
        try:
            category = Category.objects.get(pk=category_id)
        except Category.DoesNotExist:
            return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

        search_query = request.query_params.get('search', '')
        min_price = request.query_params.get('min_price')
        max_price = request.query_params.get('max_price')
        sort_by = request.query_params.get('sort_by', 'name')
        sort_order = request.query_params.get('sort_order', 'asc')

        filters = Q(visible=True, categories=category)
        if search_query:
            filters &= Q(name__icontains=search_query)
        if min_price:
            filters &= Q(discounted_price__gte=min_price)
        if max_price:
            filters &= Q(discounted_price__lte=max_price)

        try:
            products = Product.objects.filter(filters).annotate(
                effective_price=Case(
                    When(discounted_price=0, then=F('price')),
                    default=F('discounted_price'),
                    output_field=IntegerField()
                )
            )

            if sort_by == 'name':
                sort_field = 'name' if sort_order == 'asc' else '-name'
            else:
                sort_field = 'effective_price' if sort_order == 'asc' else '-effective_price'

            products = products.order_by(sort_field)

            paginator = StandardResultsSetPagination()
            paginated_products = paginator.paginate_queryset(products, request)

            serializer = ProductSerializer(paginated_products, many=True)
            return paginator.get_paginated_response(serializer.data)

        except Exception as e:
            return Response({"error": str(e)},status=status.HTTP_400_BAD_REQUEST)

class ProductsBySaleTypeView(APIView):
    def get(self, request, sale_type_id):
        try:
            sale_type = SaleType.objects.get(pk=sale_type_id)
        except SaleType.DoesNotExist:
            return Response({'error': 'SaleType not found.'}, status=status.HTTP_404_NOT_FOUND)

        products = sale_type.products.filter(visible=True).order_by('product_id')

        paginator = StandardResultsSetPagination()
        paginated_products = paginator.paginate_queryset(products, request)

        serializer = NewProductSerializer(paginated_products, many=True)

        return paginator.get_paginated_response(serializer.data)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryUpdateView(APIView):
    def put(self, request, category_id):
        try:
            category = Category.objects.get(category_id=category_id)
        except Category.DoesNotExist:
            return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

        category_data = {
            'name': request.data.get('name', category.name),
            'description': request.data.get('description', category.description),
            'visible': request.data.get('visible'),
            'home_page': request.data.get('home_page'),
            'header': request.data.get('header')
        }
        category_serializer = CategorySerializer(category, data=category_data, partial=True)

        if category_serializer.is_valid():
            category_serializer.save()
            return Response({'category': category_serializer.data}, status=status.HTTP_200_OK)

        return Response(category_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryImageUpdateView(APIView):
    def put(self, request, image_id):
        try:
            category_image = CategoryImage.objects.get(category_image_id=image_id)
        except CategoryImage.DoesNotExist:
            return Response({'error': 'Category image not found.'}, status=status.HTTP_404_NOT_FOUND)

        if 'image' in request.FILES:
            if category_image.image:
                public_id = category_image.image.public_id
                result = destroy(public_id, invalidate=True)
                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete existing image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            category_image.image = request.FILES['image']

        category_image.alt_text = request.data.get('alt_text', category_image.alt_text)
        category_image.category_id = request.data.get('category', category_image.category.category_id)
        category_image.visible = request.data.get('visible', category_image.visible)

        category_image.save()
        category_image_serializer = CategoryImageSerializer(category_image)
        return Response({'category_image': category_image_serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class ProductUpdateView(APIView):
    def put(self, request, product_id):
        try:
            product = Product.objects.get(product_id=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

        product_data = {
            'name': request.data.get('name', product.name),
            'price': request.data.get('price', product.price),
            'discounted_price': request.data.get('discounted_price', product.discounted_price),
            'short_description': request.data.get('short_description', product.short_description),
            'video_link': request.data.get('video_link', product.video_link),
            'quantity': request.data.get('quantity', product.quantity),
            'visible': request.data.get('visible', product.visible)
        }
        product_serializer = ProductSerializer(product, data=product_data, partial=True)

        if not product_serializer.is_valid():
            return Response(product_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        product_serializer.save()

        category_ids = request.data.get('categories', [])
        if isinstance(category_ids, str):
            try:
                category_ids = list(map(int, category_ids.strip("[]").split(",")))
            except ValueError:
                return Response({'error': 'Invalid categories format.'}, status=status.HTTP_400_BAD_REQUEST)

        if category_ids:
            categories = Category.objects.filter(category_id__in=category_ids)
            if categories.count() != len(category_ids):
                return Response({'error': 'One or more categories not found.'}, status=status.HTTP_400_BAD_REQUEST)
            product.categories.set(categories)

        sale_type_ids = request.data.get('sale_types', [])
        if isinstance(sale_type_ids, str):
            try:
                sale_type_ids = list(map(int, sale_type_ids.strip("[]").split(",")))
            except ValueError:
                return Response({'error': 'Invalid sale types format.'}, status=status.HTTP_400_BAD_REQUEST)

        if sale_type_ids:
            sale_types = SaleType.objects.filter(sale_type_id__in=sale_type_ids)
            if sale_types.count() != len(sale_type_ids):
                return Response({'error': 'One or more sale types not found.'}, status=status.HTTP_400_BAD_REQUEST)
            product.sale_types.set(sale_types)

        return Response({'product': product_serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class ProductImageUpdateView(APIView):
    def put(self, request, image_id):
        try:
            product_image = ProductImage.objects.get(product_image_id=image_id)
        except ProductImage.DoesNotExist:
            return Response({'error': 'Product image not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            if 'image' in request.FILES and product_image.image:
                public_id = product_image.image.public_id
                destroy(public_id, invalidate=True)
        except Exception as e:
            return Response({'error': f'Cloudinary error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


        product_image_data = {
            'product': request.data.get('product', product_image.product.product_id),
            'alt_text': request.data.get('alt_text', product_image.alt_text),
            'visible': request.data.get('visible', product_image.visible),
            'image': request.FILES.get('image') or product_image.image
        }

        serializer = ProductImageSerializer(product_image, data=product_image_data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({'product_image': serializer.data}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class FullPageContentUpdateView(APIView):
    @transaction.atomic
    def put(self, request, page_content_id):
        try:
            page_content = PageContent.objects.get(pk=page_content_id)
        except PageContent.DoesNotExist:
            return Response({'error': 'Page content not found.'}, status=status.HTTP_404_NOT_FOUND)

        page_content_data = {
            'page_name': request.data.get('page_name', page_content.page_name),
            'content': request.data.get('content', page_content.content),
            'visible': request.data.get('visible', page_content.visible)
        }
        page_content_serializer = PageContentSerializer(page_content, data=page_content_data, partial=True)

        if not page_content_serializer.is_valid():
            return Response(page_content_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        page_content_serializer.save()

        new_images = request.FILES.getlist('new_images')
        for image in new_images:
            page_image_data = {
                'page': page_content.pagecontent_id,
                'image': image,
                'visible': request.data.get('visible')
            }
            page_image_serializer = PageImageSerializer(data=page_image_data)
            if page_image_serializer.is_valid():
                page_image_serializer.save()
            else:
                transaction.set_rollback(True)
                return Response(page_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'page_content': page_content_serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class SaleTypeUpdateView(APIView):
    def put(self, request, sale_type_id):
        try:
            sale_type = SaleType.objects.get(sale_type_id=sale_type_id)
        except SaleType.DoesNotExist:
            return Response({'error': 'SaleType not found.'}, status=status.HTTP_404_NOT_FOUND)

        sale_type_data = {
            'name': request.data.get('name', sale_type.name),
            'visible': request.data.get('visible', sale_type.visible)

        }
        sale_type_serializer = SaleTypeSerializer(sale_type, data=sale_type_data, partial=True)

        if sale_type_serializer.is_valid():
            sale_type_serializer.save()
            return Response({'sale_type': sale_type_serializer.data}, status=status.HTTP_200_OK)

        return Response(sale_type_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CommentUpdateView(APIView):
    def put(self, request, comment_id):
        try:
            comment = Comment.objects.get(comment_id=comment_id)
        except Comment.DoesNotExist:
            return Response({'error': 'Comment not found.'}, status=status.HTTP_404_NOT_FOUND)

        comment_data = {
            'product': request.data.get('product', comment.product.product_id),
            'user_name': request.data.get('user_name', comment.user_name),
            'rating': request.data.get('rating', comment.rating),
            'text': request.data.get('text', comment.text),
            'display': request.data.get('display', comment.display)
        }
        comment_serializer = CommentSerializer(comment, data=comment_data, partial=True)

        if comment_serializer.is_valid():
            comment_serializer.save()
            return Response({'comment': comment_serializer.data}, status=status.HTTP_200_OK)

        return Response(comment_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class BannerImageUpdateView(APIView):
    def put(self, request, banner_image_id):
        try:
            banner_image = BannerImage.objects.get(banner_image_id=banner_image_id)
        except BannerImage.DoesNotExist:
            return Response({'error': 'BannerImage not found.'}, status=status.HTTP_404_NOT_FOUND)

        if 'image' in request.FILES and banner_image.image:
            public_id = banner_image.image.public_id
            destroy(public_id, invalidate=True)

        banner_image_data = {
            'title': request.data.get('title', banner_image.title),
            'visible': request.data.get('visible', banner_image.visible),
            'image': request.FILES.get('image') or banner_image.image
        }

        serializer = BannerImageSerializer(banner_image, data=banner_image_data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({'banner_image': serializer.data}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class CustomerUpdateView(APIView):
    def put(self, request, customer_id):
        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            return Response({'error': 'Customer not found.'}, status=status.HTTP_404_NOT_FOUND)

        customer_data = {
            'name': request.data.get('name', customer.name),
            'phone_number': request.data.get('phone_number', customer.phone_number),
            'email': request.data.get('email', customer.email),
            'address': request.data.get('address', customer.address),
            'pincode': request.data.get('pincode', customer.pincode),
            'visible': request.data.get('visible', customer.visible),
        }
        customer_serializer = CustomerSerializer(customer, data=customer_data, partial=True)

        if customer_serializer.is_valid():
            customer_serializer.save()
            return Response({'customer': customer_serializer.data}, status=status.HTTP_200_OK)

        return Response(customer_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class OrderUpdateView(APIView):
    def put(self, request, order_id):
        try:
            order = Order.objects.get(order_id=order_id)
        except Order.DoesNotExist:
            return Response({'error': 'Order not found.'}, status=status.HTTP_404_NOT_FOUND)

        order_data = {
            'total_amount': request.data.get('total_amount', order.total_amount),
            'count': request.data.get('count', order.count),
        }
        order_serializer = OrderSerializer(order, data=order_data, partial=True)

        if order_serializer.is_valid():
            order_serializer.save()
            return Response({'order': order_serializer.data}, status=status.HTTP_200_OK)

        return Response(order_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class OrderItemUpdateView(APIView):
    def put(self, request, order_item_id):
        try:
            order_item = OrderItem.objects.get(order_item_id=order_item_id)
        except OrderItem.DoesNotExist:
            return Response({'error': 'OrderItem not found.'}, status=status.HTTP_404_NOT_FOUND)

        order_item_data = {
            'quantity': request.data.get('quantity', order_item.quantity),
            'price': request.data.get('price', order_item.price),
            'visible': request.data.get('visible', order_item.visible),
        }
        order_item_serializer = OrderItemSerializer(order_item, data=order_item_data, partial=True)

        if order_item_serializer.is_valid():
            order_item_serializer.save()
            return Response({'order_item': order_item_serializer.data}, status=status.HTTP_200_OK)

        return Response(order_item_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class UpdateOrderView(APIView):
    @transaction.atomic
    def put(self, request):
        try:
            data = json.loads(request.body)
            order_id = data.get('order_id')
            order_items_data = data.get('items')
            note = data.get('note')

            order = Order.objects.get(order_id=order_id)

            total_amount = sum(item.price for item in order.items.all())
            total_count = sum(item.quantity for item in order.items.all())

            for item_data in order_items_data:
                product_id = item_data.get('product_id')
                new_quantity = item_data.get('quantity')

                if not product_id or not new_quantity:
                    return Response({'error': 'Invalid product or quantity.'}, status=status.HTTP_400_BAD_REQUEST)

                try:
                    product = Product.objects.get(product_id=product_id)
                except Product.DoesNotExist:
                    return Response({'error': f"Product {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)

                if new_quantity > product.quantity:
                    return Response({
                        'error': f"Insufficient stock for product {product.name}. Available quantity: {product.quantity}"
                    }, status=status.HTTP_400_BAD_REQUEST)

                try:
                    order_item = OrderItem.objects.get(order=order, product_id=product_id)

                    total_amount -= order_item.price
                    total_count -= order_item.quantity

                    price_to_use = product.discounted_price if product.discounted_price > 0 else product.price

                    order_item.quantity = new_quantity
                    order_item.price = price_to_use * new_quantity
                    order_item.order_completed = True
                    order_item.save()

                    total_amount += order_item.price
                    total_count += order_item.quantity
                except OrderItem.DoesNotExist:
                    return Response({'error': f"Order item for product {product_id} not found."}, status=status.HTTP_404_NOT_FOUND)

            order.count = total_count
            order.total_amount = total_amount
            order.note = note

            if all(item.order_completed for item in order.items.all()):
                order.order_completed = True
            else:
                order.order_completed = False

            order.save()

            return Response({'message': 'Order and items updated successfully.'}, status=status.HTTP_200_OK)

        except Order.DoesNotExist:
            return Response({'error': 'Order not found.'}, status=status.HTTP_404_NOT_FOUND)
        except json.JSONDecodeError:
            return Response({'error': 'Invalid JSON data.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            transaction.set_rollback(True)
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopUpdateView(APIView):
    @transaction.atomic
    def put(self, request, workshop_id):
        try:
            workshop = Workshop.objects.get(workshop_id=workshop_id)
        except Workshop.DoesNotExist:
            return Response({'error': 'Workshop not found.'}, status=status.HTTP_404_NOT_FOUND)

        workshop_data = {
            'name': request.data.get('name', workshop.name),
            'date': request.data.get('date', workshop.date),
            'place': request.data.get('place', workshop.place),
            'description': request.data.get('description', workshop.description),
            'completed': request.data.get('completed', workshop.completed)
        }

        workshop_serializer = WorkshopSerializer(workshop, data=workshop_data, partial=True)
        if not workshop_serializer.is_valid():
            return Response(workshop_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        workshop_serializer.save()

        videos_data = request.data.get('videos', '[]')
        if isinstance(videos_data, str):
            try:
                videos_data = json.loads(videos_data)
            except json.JSONDecodeError:
                return Response({'error': 'Invalid JSON format for videos.'}, status=status.HTTP_400_BAD_REQUEST)

        if not isinstance(videos_data, list):
            return Response({'error': 'Videos data should be a list.'}, status=status.HTTP_400_BAD_REQUEST)

        for video_data in videos_data:
            video_id = video_data.get('video_id')
            new_video_url = video_data.get('video_url')

            if not video_id or not new_video_url:
                continue

            try:
                workshop_video = WorkshopVideo.objects.get(workshopvideo_id=video_id, workshop=workshop)
                workshop_video.video_url = new_video_url
                workshop_video.save()
            except WorkshopVideo.DoesNotExist:
                return Response({'error': f'Workshop video with ID {video_id} not found.'}, status=status.HTTP_404_NOT_FOUND)

        return Response({'workshop': workshop_serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class AddCategoryImageView(APIView):
    def post(self, request, category_id):
        try:
            category = Category.objects.get(pk=category_id)
        except Category.DoesNotExist:
            return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

        image = request.FILES.get('image')
        alt_text = request.data.get('alt_text', '')

        if not image:
            return Response({'error': 'No image file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        category_image_data = {
            'category': category_id,
            'image': image,
            'alt_text': alt_text
        }
        category_image_serializer = CategoryImageSerializer(data=category_image_data)

        if category_image_serializer.is_valid():
            category_image_serializer.save()
            return Response({'category_image': category_image_serializer.data}, status=status.HTTP_201_CREATED)
        else:
            return Response(category_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class AddProductImageView(APIView):
    def post(self, request, product_id):
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

        image = request.FILES.get('image')
        alt_text = request.data.get('alt_text', '')

        if not image:
            return Response({'error': 'No image file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        product_image_data = {
            'product': product_id,
            'image': image,
            'alt_text': alt_text
        }
        product_image_serializer = ProductImageSerializer(data=product_image_data)

        if product_image_serializer.is_valid():
            product_image_serializer.save()
            return Response({'product_image': product_image_serializer.data}, status=status.HTTP_201_CREATED)
        else:
            return Response(product_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class AddPageImageView(APIView):
    @transaction.atomic
    def post(self, request, page_content_id):
        try:
            page_content = PageContent.objects.get(pk=page_content_id)
        except PageContent.DoesNotExist:
            return Response({'error': 'Page content not found.'}, status=status.HTTP_404_NOT_FOUND)

        page_image = request.FILES.get('image')

        if not page_image:
            return Response({'error': 'No image file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        page_image_instance = PageImage(page=page_content, image=page_image)
        page_image_instance.save()

        return Response({
            'message': 'Page image uploaded successfully.',
            'page_image': {
                'page_image_id': page_image_instance.pageimage_id,
                'page_name': page_content.page_name,
                'image_url': page_image_instance.image.url
            }
        }, status=status.HTTP_201_CREATED)

@method_decorator(csrf_exempt, name='dispatch')
class SaleTypeDeleteView(APIView):
    def delete(self, request, sale_type_id):
        try:
            sale_type = SaleType.objects.get(pk=sale_type_id)
            sale_type.delete()
            return Response({'message': 'SaleType deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except SaleType.DoesNotExist:
            return Response({'error': 'SaleType not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryDeleteView(APIView):
    def delete(self, request, category_id):
        try:
            category = Category.objects.get(pk=category_id)
            category.delete()
            return Response({'message': 'Category deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except Category.DoesNotExist:
            return Response({'error': 'Category not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class ProductDeleteView(APIView):
    def delete(self, request, product_id):
        try:
            product = Product.objects.get(pk=product_id)
            product.delete()
            return Response({'message': 'Product deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except Product.DoesNotExist:
            return Response({'error': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class CategoryImageDeleteView(APIView):
    def delete(self, request, category_image_id):
        try:
            category_image = CategoryImage.objects.get(pk=category_image_id)
            if category_image.image:
                public_id = category_image.image.public_id
                result = destroy(public_id, invalidate=True)

                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            category_image.delete()
            return Response({'message': 'Category image deleted successfully from both database and Cloudinary.'}, status=status.HTTP_204_NO_CONTENT)

        except CategoryImage.DoesNotExist:
            return Response({'error': 'Category image not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class ProductImageDeleteView(APIView):
    def delete(self, request, product_image_id):
        try:
            product_image = ProductImage.objects.get(pk=product_image_id)
            if product_image.image:
                public_id = product_image.image.public_id
                result = destroy(public_id, invalidate=True)
                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            product_image.delete()
            return Response({'message': 'Product image deleted successfully from both database and Cloudinary.'}, status=status.HTTP_204_NO_CONTENT)
        except ProductImage.DoesNotExist:
            return Response({'error': 'Product image not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class CommentDeleteView(APIView):
    def delete(self, request, comment_id):
        try:
            comment = Comment.objects.get(pk=comment_id)
            comment.delete()
            return Response({'message': 'Comment deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except Comment.DoesNotExist:
            return Response({'error': 'Comment not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class BannerImageDeleteView(APIView):
    def delete(self, request, banner_image_id):
        try:
            banner_image = BannerImage.objects.get(pk=banner_image_id)
            if banner_image.image:
                public_id = banner_image.image.public_id
                result = destroy(public_id, invalidate=True)
                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            banner_image.delete()
            return Response({'message': 'Banner image deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except BannerImage.DoesNotExist:
            return Response({'error': 'Banner image not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class CustomerDeleteView(APIView):
    def delete(self, request, customer_id):
        try:
            customer = Customer.objects.get(pk=customer_id)
            customer.delete()
            return Response({'message': 'Customer deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except Customer.DoesNotExist:
            return Response({'error': 'Customer not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class OrderDeleteView(APIView):
    def delete(self, request, order_id):
        try:
            order = Order.objects.get(pk=order_id)
            order.delete()
            return Response({'message': 'Order deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except Order.DoesNotExist:
            return Response({'error': 'Order not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class OrderItemDeleteView(APIView):
    def delete(self, request, order_item_id):
        try:
            order_item = OrderItem.objects.get(pk=order_item_id)
            order_item.delete()
            return Response({'message': 'Order item deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except OrderItem.DoesNotExist:
            return Response({'error': 'Order item not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class PageContentDeleteView(APIView):
    def delete(self, request, page_content_id):
        try:
            page_content = PageContent.objects.get(pk=page_content_id)
            page_content.delete()
            return Response({'message': 'Page content and associated images deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except PageContent.DoesNotExist:
            return Response({'error': 'Page content not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class PageImageDeleteView(APIView):
    def delete(self, request, page_content_id):
        try:
            page_content = PageContent.objects.get(pk=page_content_id)
            page_images = PageImage.objects.filter(page=page_content)
            if page_images.exists():
                for image in page_images:
                    if image.image:
                        public_id = image.image.public_id
                        result = destroy(public_id, invalidate=True)
                        if result.get('result') != 'ok':
                            return Response({'error': 'Failed to delete an image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                page_images.delete()
                return Response({'message': 'Images for the specified page content deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
            else:
                return Response({'message': 'No images found for the specified page content.'}, status=status.HTTP_404_NOT_FOUND)
        except PageContent.DoesNotExist:
            return Response({'error': 'Page content not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopDeleteView(APIView):
    def delete(self, request, workshop_id):
        try:
            workshop = Workshop.objects.get(pk=workshop_id)
            workshop.delete()
            return Response({'message': 'Workshop deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except Workshop.DoesNotExist:
            return Response({'error': 'Workshop not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopImageDeleteView(APIView):
    def delete(self, request, workshop_image_id):
        try:
            workshop_image = WorkshopImage.objects.get(pk=workshop_image_id)
            if workshop_image.image:
                public_id = workshop_image.image.public_id
                result = destroy(public_id, invalidate=True)
                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            workshop_image.delete()
            return Response({'message': 'Workshop image deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except WorkshopImage.DoesNotExist:
            return Response({'error': 'Workshop image not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopVideoDeleteView(APIView):
    def delete(self, request, workshop_video_id):
        try:
            workshop_video = WorkshopVideo.objects.get(pk=workshop_video_id)
            workshop_video.delete()
            return Response({'message': 'Workshop video deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except WorkshopVideo.DoesNotExist:
            return Response({'error': 'Workshop video not found.'}, status=status.HTTP_404_NOT_FOUND)

@csrf_exempt
def send_message_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            message_body = data.get('message', '')
            to_number = '917353647516'  # Replace with the actual recipient number

            formatted_message = f"{message_body}"

            message_sid = send_whatsapp_message(to_number, formatted_message)
            return JsonResponse({'status': 'success', 'message_sid': message_sid})
        except Exception as e:
            return JsonResponse({'status': 'error', 'error': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'error': 'Invalid method'}, status=405)

class LoginView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return Response({'message': 'Logged in successfully'}, status=status.HTTP_200_OK)
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

class LogoutView(APIView):
    def post(self, request):
        logout(request)
        return Response({'message': 'Logged out successfully'}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopImageUpdateView(APIView):
    def put(self, request, image_id):
        try:
            workshop_image = WorkshopImage.objects.get(workshopimage_id=image_id)
        except WorkshopImage.DoesNotExist:
            return Response({'error': 'Workshop image not found.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            if 'image' in request.FILES and workshop_image.image:
                public_id = workshop_image.image.public_id
                destroy(public_id, invalidate=True)
        except Exception as e:
            return Response({'error': f'Cloudinary error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        workshop_image_data = {
            'workshop': request.data.get('workshop', workshop_image.workshop.workshop_id),
            'visible': request.data.get('visible', workshop_image.visible),
            'image': request.FILES.get('image') or workshop_image.image
        }

        serializer = WorkshopImageSerializer(workshop_image, data=workshop_image_data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({'workshop_image': serializer.data}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopCreateView(APIView):
    @transaction.atomic
    def post(self, request):
        workshop_data = {
            'name': request.data.get('name'),
            'date': request.data.get('date'),
            'place': request.data.get('place'),
            'description': request.data.get('description'),
            'completed': request.data.get('completed', False),
        }

        video_url = request.data.get('video_url')
        images_data = request.FILES.getlist('images')

        workshop_serializer = WorkshopSerializer(data=workshop_data)
        if workshop_serializer.is_valid():
            workshop = workshop_serializer.save()

            if video_url:
                video_data = {'workshop': workshop.workshop_id, 'video_url': video_url}  # Pass the Workshop instance
                video_serializer = AddWorkshopVideoSerializer(data=video_data)
                if video_serializer.is_valid():
                    video_serializer.save()
                else:
                    transaction.set_rollback(True)
                    return Response(video_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            for image in images_data:
                image_data = {'workshop': workshop.workshop_id, 'image': image}  # Pass the Workshop instance
                image_serializer = AddWorkshopImageSerializer(data=image_data)
                if image_serializer.is_valid():
                    image_serializer.save()

            return Response({'message': 'Workshop, video, and images saved successfully','workshop': workshop_serializer.data,}, status=status.HTTP_201_CREATED)

        return Response(workshop_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopImageView(APIView):
    def post(self, request, workshop_id):
        try:
            workshop = Workshop.objects.get(pk=workshop_id)
        except Workshop.DoesNotExist:
            return Response({'error': 'Workshop not found.'}, status=status.HTTP_404_NOT_FOUND)

        images = request.FILES.get('image')
        if not images:
            return Response({'error': 'No images provided.'}, status=status.HTTP_400_BAD_REQUEST)

        workshop_image_data = {'workshop': workshop.pk, 'image': images}
        image_serializer = AddWorkshopImageSerializer(data=workshop_image_data)
        if image_serializer.is_valid():
            image_serializer.save()
        else:
            return Response(image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'Images added successfully.'}, status=status.HTTP_201_CREATED)

@method_decorator(csrf_exempt, name='dispatch')
class WorkshopVideoView(APIView):
    def post(self, request, workshop_id):
        try:
            workshop = Workshop.objects.get(pk=workshop_id)
        except Workshop.DoesNotExist:
            return Response({'error': 'Workshop not found.'}, status=status.HTTP_404_NOT_FOUND)

        video_urls = request.data.getlist('video_urls')
        if not video_urls:
            return Response({'error': 'No video URLs provided.'}, status=status.HTTP_400_BAD_REQUEST)

        for video_url in video_urls:
            video_data = {'workshop': workshop.pk, 'video_url': video_url}
            video_serializer = AddWorkshopVideoSerializer(data=video_data)
            if video_serializer.is_valid():
                video_serializer.save()
            else:
                return Response(video_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({'message': 'Videos added successfully.'}, status=status.HTTP_201_CREATED)

def send_order_email(customer, order, order_items):
    try:
        # Collect product images
        order_items_with_images = []
        for item in order_items:
            first_image = item.product.images.first()  # Get the first image of the product
            image_url = (
                f"https://res.cloudinary.com/dgkgxokru/{first_image.image}" if first_image else None
            )
            order_items_with_images.append({
                'product': item.product,
                'quantity': item.quantity,
                'price': item.price,
                'image_url': image_url,
            })

        # Render email HTML content
        html_message = render_to_string('kalegooduapp/order_details.html', {
            'customer': customer,
            'order': order,
            'order_items': order_items_with_images,
        })

        # Create and send email
        email = EmailMultiAlternatives(
            subject="Order Confirmation",
            body="Thank you for your order. Please find the details attached.",
            from_email="photo2pruthvi@gmail.com",
            to=[customer.email],
        )
        email.attach_alternative(html_message, "text/html")
        email.send()

        return True
    except Exception as e:
        print(f"Error sending email: {e}")
        return False

# @method_decorator(csrf_exempt, name='dispatch')
# class CreateOrderView(APIView):
#     @transaction.atomic
#     def post(self, request):
#         try:
#             customer_data = request.data.get('customerDetails', '{}')
#             order_data = request.data.get('orderDetails', '{}')
#             items_data = order_data.get('items', '[]')

#             # Save customer data
#             customer_serializer = CustomerSerializer(data=customer_data)
#             if customer_serializer.is_valid():
#                 customer = customer_serializer.save()
#             else:
#                 return Response(customer_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#             # Save order data
#             order_data = {
#                 'customer': customer.customer_id,
#                 'total_amount': order_data['total'],
#                 'count': order_data['count'],
#             }
#             order_serializer = OrderSerializer(data=order_data)
#             if order_serializer.is_valid():
#                 order = order_serializer.save()
#             else:
#                 transaction.set_rollback(True)
#                 return Response(order_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#             # Save order items
#             for item in items_data:
#                 item_data = {
#                     'order': order.order_id,
#                     'product': item['product_id'],
#                     'quantity': item['quantity'],
#                     'price': item['price'],
#                 }
#                 order_item_serializer = OrderItemSerializer(data=item_data)
#                 if order_item_serializer.is_valid():
#                     order_item_serializer.save()
#                 else:
#                     transaction.set_rollback(True)
#                     return Response(order_item_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#             # Fetch all order items
#             order_items = order.items.all()

#             # Send email to customer
#             email_sent = send_order_email(customer, order, order_items)
#             if not email_sent:
#                 return Response({'error': 'Order created, but email failed to send.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#             return Response({'message': 'Order created and email sent successfully.'}, status=status.HTTP_201_CREATED)

#         except Exception as e:
#             transaction.set_rollback(True)
#             return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@method_decorator(csrf_exempt, name='dispatch')
class CreateOrderView(APIView):
    @transaction.atomic
    def post(self, request):
        try:
            customer_data = request.data.get('customerDetails', '{}')
            order_data = request.data.get('orderDetails', '{}')
            items_data = order_data.get('items', '[]')

            # Save customer data
            customer_serializer = CustomerSerializer(data=customer_data)
            if customer_serializer.is_valid():
                customer = customer_serializer.save()
            else:
                return Response(customer_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            # Save order data
            order_data = {
                'customer': customer.customer_id,
                'total_amount': order_data['total'],
                'count': order_data['count'],
            }
            order_serializer = OrderSerializer(data=order_data)
            if order_serializer.is_valid():
                order = order_serializer.save()
            else:
                transaction.set_rollback(True)
                return Response(order_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            # Save order items and update product quantities
            for item in items_data:
                try:
                    # Fetch product and check stock availability
                    product = Product.objects.select_for_update().get(pk=item['product_id'])
                    if product.quantity < item['quantity']:
                        transaction.set_rollback(True)
                        return Response(
                            {'error': f"Insufficient stock for product '{product.name}'. Available: {product.quantity}, Requested: {item['quantity']}."},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    # Prepare order item data
                    item_data = {
                        'order': order.order_id,
                        'product': product.product_id,
                        'quantity': item['quantity'],
                        'price': item['price'],
                    }

                    # Validate and save order item
                    order_item_serializer = OrderItemSerializer(data=item_data)
                    if order_item_serializer.is_valid():
                        order_item_serializer.save()
                        print(order_item_serializer)

                        # Update product quantity
                        product.quantity -= item['quantity']
                        product.save()
                    else:
                        transaction.set_rollback(True)
                        return Response(order_item_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

                except Product.DoesNotExist:
                    transaction.set_rollback(True)
                    return Response(
                        {'error': f"Product with ID {item['product_id']} does not exist."},
                        status=status.HTTP_404_NOT_FOUND,
                    )
                except Exception as e:
                    transaction.set_rollback(True)
                    return Response(
                        {'error': f"Error processing order item: {str(e)}"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

            # Fetch all order items
            order_items = order.items.all()

            # Send email to customer
            email_sent = send_order_email(customer, order, order_items)
            if not email_sent:
                return Response({'error': 'Order created, but email failed to send.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            return Response({'message': 'Order created and email sent successfully.'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            transaction.set_rollback(True)
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportCustomersOrdersView(View):
    def get(self, request):
        workbook = Workbook()

        customer_sheet = workbook.active
        customer_sheet.title = "Customers"
        customer_sheet.append([
            'Customer ID', 'Customer Name', 'Phone Number', 'Email', 'Address', 'Pincode', 'Created At', 'Updated At'
        ])
        customers = Customer.objects.all()
        for customer in customers:
            customer_sheet.append([
                customer.customer_id,
                customer.name,
                customer.phone_number,
                customer.email,
                customer.address,
                customer.pincode,
                customer.created_at.replace(tzinfo=None) if customer.created_at else None,
                customer.updated_at.replace(tzinfo=None) if customer.updated_at else None,
            ])

        order_sheet = workbook.create_sheet(title="Orders")
        order_sheet.append([
            'Order ID', 'Customer Name', 'Total Amount', 'Order Count', 'Order Completed', 'Created At', 'Updated At'
        ])
        orders = Order.objects.select_related('customer').all()
        for order in orders:
            order_sheet.append([
                order.order_id,
                order.customer.name,
                order.total_amount,
                order.count,
                order.order_completed,
                order.created_at.replace(tzinfo=None) if order.created_at else None,
                order.updated_at.replace(tzinfo=None) if order.updated_at else None,
            ])

        detailed_sheet = workbook.create_sheet(title="Order Details")
        detailed_sheet.append([
            'Order ID', 'Customer Name', 'Product Name', 'Quantity', 'Price', 'Created At', 'Updated At'
        ])
        order_items = OrderItem.objects.select_related('order', 'product', 'order__customer').all()
        for item in order_items:
            detailed_sheet.append([
                item.order.order_id,
                item.order.customer.name,
                item.product.name,
                item.quantity,
                item.price,
                item.created_at.replace(tzinfo=None) if item.created_at else None,
                item.updated_at.replace(tzinfo=None) if item.updated_at else None,
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Customers_Orders.xlsx"'
        workbook.save(response)

        return response

class ExportCustomersOrdersByDateView(View):
    def get(self, request):
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        print(start_date_str)
        print(end_date_str)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        except ValueError:
            return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

        workbook = Workbook()
        print(start_date)
        print(end_date)

        customer_sheet = workbook.active
        customer_sheet.title = "Customers"
        customer_sheet.append([
            'Customer ID', 'Customer Name', 'Phone Number', 'Email', 'Address', 'Pincode', 'Created At', 'Updated At'
        ])
        customers = Customer.objects.all()
        if start_date and end_date:
            customers = customers.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        for customer in customers:
            customer_sheet.append([
                customer.customer_id,
                customer.name,
                customer.phone_number,
                customer.email,
                customer.address,
                customer.pincode,
                customer.created_at.replace(tzinfo=None) if customer.created_at else None,
                customer.updated_at.replace(tzinfo=None) if customer.updated_at else None,
            ])

        order_sheet = workbook.create_sheet(title="Orders")
        order_sheet.append([
            'Order ID', 'Customer Name', 'Total Amount', 'Order Count', 'Order Completed', 'Created At', 'Updated At'
        ])
        orders = Order.objects.select_related('customer').all()
        if start_date and end_date:
            orders = orders.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

        for order in orders:
            order_sheet.append([
                order.order_id,
                order.customer.name,
                order.total_amount,
                order.count,
                order.order_completed,
                order.created_at.replace(tzinfo=None) if order.created_at else None,
                order.updated_at.replace(tzinfo=None) if order.updated_at else None,
            ])

        detailed_sheet = workbook.create_sheet(title="Order Details")
        detailed_sheet.append([
            'Order ID', 'Customer Name', 'Product Name', 'Quantity', 'Price', 'Created At', 'Updated At'
        ])
        order_items = OrderItem.objects.select_related('order', 'product', 'order__customer').all()
        if start_date and end_date:
            order_items = order_items.filter(order__created_at__date__gte=start_date, order__created_at__date__lte=end_date)

        for item in order_items:
            detailed_sheet.append([
                item.order.order_id,
                item.order.customer.name,
                item.product.name,
                item.quantity,
                item.price,
                item.created_at.replace(tzinfo=None) if item.created_at else None,
                item.updated_at.replace(tzinfo=None) if item.updated_at else None,
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Customers_Orders_{start_date_str}_to_{end_date_str}.xlsx"'
        workbook.save(response)

        return response

# Authorize Razorpay client with API Keys
razorpay_client = razorpay.Client(
    auth=(settings.RAZOR_KEY_ID, settings.RAZORPAY_API_SECRET)
)

@csrf_exempt
def create_order(request):
    if request.method == "POST":
        try:
            # Parse the JSON body from the frontend
            data = json.loads(request.body)
            print(data)
            amount = data.get("amount", 100)  # Default to Rs. 200
            currency = data.get("currency", "INR")

            # Create a Razorpay Order
            razorpay_order = razorpay_client.order.create(
                dict(amount=amount, currency=currency, payment_capture="0")
            )

            # Prepare response with order details
            response_data = {
                "razorpay_order_id": razorpay_order["id"],
                "razorpay_merchant_key": settings.RAZOR_KEY_ID,
                "amount": amount,
                "currency": currency,
            }
            return JsonResponse(response_data, status=200)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return HttpResponseBadRequest("Invalid request method")

@csrf_exempt
def paymenthandler(request):
    if request.method == "POST":
        try:
            # Parse the JSON body from the frontend
            data = json.loads(request.body)
            print("Hiii")
            print(data)
            payment_id = data.get("razorpay_payment_id", "")
            razorpay_order_id = data.get("razorpay_order_id", "")
            signature = data.get("razorpay_signature", "")

            # Prepare the parameters for verification
            params_dict = {
                "razorpay_order_id": razorpay_order_id,
                "razorpay_payment_id": payment_id,
                "razorpay_signature": signature,
            }

            # Verify the payment signature
            result = razorpay_client.utility.verify_payment_signature(params_dict)
            if result is not None:
                try:
                    # Capture the payment
                    amount = data.get("amount", 100)  # Amount in paise
                    razorpay_client.payment.capture(payment_id, amount)

                    # Respond with success status
                    return JsonResponse({"status": "success"}, status=200)
                except Exception as e:
                    return JsonResponse({"error": str(e)}, status=400)
            else:
                return JsonResponse({"status": "signature_verification_failed"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return HttpResponseBadRequest("Invalid request method")

@method_decorator(csrf_exempt, name='dispatch')
class SendProductPromotionalEmails(APIView):
    def post(self, request):
        try:
            data = request.data

            title = data.get("title")
            body = data.get("body")
            product_ids = data.get("products", [])

            if not title or not body or not product_ids:
                return JsonResponse({"error": "Missing required fields (title, body, or products)"}, status=400)

            # Fetch products based on the product IDs
            products = Product.objects.filter(product_id__in=product_ids, visible=True)

            if not products.exists():
                return JsonResponse({"error": "No products found for the given product IDs."}, status=404)

            # Prepare product data with image URL and price
            product_info = []
            for product in products:
                product_image = product.images.first()  # Get the first image associated with the product
                image_url = f"https://res.cloudinary.com/dgkgxokru/{product_image.image}" if product_image else None

                product_info.append({
                    'name': product.name,
                    'price': product.discounted_price if product.discounted_price > 0 else product.price,
                    'image_url': image_url,
                })

            # Loop over each customer and send the promotional email
            customers = Customer.objects.filter(send_promotion_emails=True).values('email').distinct()

            for customer in customers:
                try:
                    # Prepare unsubscribe link
                    unsubscribe_link = f"https://kalegoodupractice.pythonanywhere.com/api/unsubscribe/{customer['email']}/"

                    # Render the promotional email content with unsubscribe link
                    html_message = render_to_string('kalegooduapp/promotional_email.html', {
                        'message': body,  # Include the body content from the request
                        'title': title,   # Include the title content from the request
                        'unsubscribe_link': unsubscribe_link,
                        'products': product_info,  # Pass the products and their info to the template
                    })

                    # Create and send the email
                    email = EmailMultiAlternatives(
                        subject=title,  # Use the provided title for the email subject
                        body="Check out our new products! For more details, see the attached HTML.",
                        from_email="photo2pruthvi@gmail.com",
                        to=[customer['email']],
                    )
                    email.attach_alternative(html_message, "text/html")
                    email.send()

                    print(f"Promotional email sent to {customer['email']}")
                except Exception as e:
                    print(f"Error sending email to {customer['email']}: {e}")
            return JsonResponse({"message": "Promotional emails sent successfully to all customers."}, status=200)

        except Exception as e:
            print(f"Error sending emails: {e}")
            return JsonResponse({"error": "Error sending emails. Please try again later."}, status=500)

class UnsubscribeView(View):
    def get(self, request, email):
        try:
            customers = Customer.objects.filter(email=email)

            if not customers.exists():
                return HttpResponse("No customer found with the provided email.", status=404)

            # Update all matching records
            customers.update(send_promotion_emails=False)

            return HttpResponse("You have successfully unsubscribed from promotional emails.", status=200)

        except Exception as e:
            return HttpResponse(f"An error occurred: {str(e)}", status=500)

class ProductProductIdView(APIView):
    def get(self, request):
        try:
            products = Product.objects.all()  # Get all products
            serializer = ProductTestimonialSerializer(products, many=True)  # Serialize the products
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@method_decorator(csrf_exempt, name='dispatch')
class SendWorkshopPromotionEmails(View):
    def post(self, request):
        try:
            data = request.data

            title = data.get("title")
            body = data.get("body")
            workshop_ids = data.get("workshops", [])

            if not title or not body or not workshop_ids:
                return JsonResponse({"error": "Missing required fields (title, body, or workshops)"}, status=400)

            # Fetch workshops based on provided IDs
            workshops = Workshop.objects.filter(workshop_id__in=workshop_ids)

            if not workshops.exists():
                return JsonResponse({"error": "No upcoming workshops found for the given IDs."}, status=404)

            # Prepare workshop information
            workshop_info = []
            for workshop in workshops:
                workshop_image = workshop.images.first()  # Assuming related images exist
                image_url = f"https://res.cloudinary.com/dgkgxokru/{workshop_image.image}" if workshop_image else None

                workshop_info.append({
                    'name': workshop.name,
                    # 'date': workshop.date,
                    'place': workshop.place,
                    'description': workshop.description,
                    'image_url': image_url,
                })

            # Fetch customers who have not unsubscribed
            customers = Customer.objects.filter(send_promotion_emails=True).values('email').distinct()

            for customer in customers:
                try:
                    # Prepare unsubscribe link
                    unsubscribe_link = f"https://kalegoodupractice.pythonanywhere.com/api/unsubscribe/{customer['email']}/"

                    # Render email content
                    html_message = render_to_string('kalegooduapp/workshop_promotion_email.html', {
                        'message': body,
                        'title': title,
                        'unsubscribe_link': unsubscribe_link,
                        'workshops': workshop_info,
                    })

                    # Send email
                    email = EmailMultiAlternatives(
                        subject=title,
                        body="Check out our upcoming workshops!",
                        from_email="photo2pruthvi@gmail.com",
                        to=[customer['email']],
                    )
                    email.attach_alternative(html_message, "text/html")
                    email.send()

                    print(f"Workshop promotion email sent to {customer['email']}")
                except Exception as e:
                    print(f"Error sending email to {customer['email']}: {e}")

            return JsonResponse({"message": "Workshop promotion emails sent successfully!"}, status=200)

        except Exception as e:
            print(f"Error sending emails: {e}")
            return JsonResponse({"error": "Error sending emails. Please try again later."}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class ValidateStockView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            cart_items = data.get('cartItems', [])

            out_of_stock = []
            for item in cart_items:
                try:
                    product = Product.objects.get(product_id=item['product_id'])
                    if product.quantity < int(item['cartQuantity']):
                        out_of_stock.append(product.name)
                except Product.DoesNotExist:
                    out_of_stock.append(f"Product with ID {item['product_id']} not found")

            return JsonResponse({'valid': len(out_of_stock) == 0, 'outOfStockItems': out_of_stock}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON format'}, status=400)
        except Exception as e:
            return JsonResponse({'error': f'Error validating stock: {str(e)}'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class SubCategoryCreateView(APIView):
    @transaction.atomic
    def post(self, request):
        categories_data = request.data.get('categories', '[]') 
        if isinstance(categories_data, str):
            try:
                categories_data = json.loads(categories_data)
            except json.JSONDecodeError:
                return Response({'error': 'Invalid data format for categories.'}, status=status.HTTP_400_BAD_REQUEST)

        if not categories_data or not isinstance(categories_data, list):
            return Response({'category': ['This field is required.']}, status=status.HTTP_400_BAD_REQUEST)

        categories = Category.objects.filter(category_id__in=categories_data)
        if not categories.exists():
            return Response({'error': 'One or more categories do not exist.'}, status=status.HTTP_400_BAD_REQUEST)

        subcategory_data = {
            'name': request.data.get('name'),
            'description': request.data.get('description'),
            'visible': request.data.get('visible', True),
            'header': request.data.get('header', False),
            'category': [category.category_id for category in categories],
            'category_page': request.data.get('category_page', False),
        }

        subcategory_serializer = SubCategorySerializer(data=subcategory_data)
        if subcategory_serializer.is_valid():
            subcategory = subcategory_serializer.save()
            subcategory.category.set(categories)

            # Handling image uploads
            images = request.FILES.getlist('images')
            for image in images:
                subcategory_image_data = {
                    'subcategory': subcategory.subcategory_id,
                    'image': image,
                    'alt_text': request.data.get('alt_text', ''),
                    'visible': request.data.get('visible')
                }
                subcategory_image_serializer = SubCategoryImageSerializer(data=subcategory_image_data)
                if subcategory_image_serializer.is_valid():
                    subcategory_image_serializer.save()
                else:
                    transaction.set_rollback(True)
                    return Response(subcategory_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'subcategory': subcategory_serializer.data}, status=status.HTTP_201_CREATED)

        return Response(subcategory_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class AddSubCategoryImageView(APIView):
    def post(self, request, subcategory_id):
        try:
            subcategory = SubCategory.objects.get(pk=subcategory_id)
        except SubCategory.DoesNotExist:
            return Response({'error': 'SubCategory not found.'}, status=status.HTTP_404_NOT_FOUND)

        image = request.FILES.get('image')
        alt_text = request.data.get('alt_text', '')

        if not image:
            return Response({'error': 'No image file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        subcategory_image_data = {
            'subcategory': subcategory_id,
            'image': image,
            'alt_text': alt_text
        }
        subcategory_image_serializer = SubCategoryImageSerializer(data=subcategory_image_data)

        if subcategory_image_serializer.is_valid():
            subcategory_image_serializer.save()
            return Response({'subcategory_image': subcategory_image_serializer.data}, status=status.HTTP_201_CREATED)
        else:
            return Response(subcategory_image_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@method_decorator(csrf_exempt, name='dispatch')
class SubCategoryUpdateView(APIView):
    def put(self, request, subcategory_id):
        try:
            subcategory = SubCategory.objects.get(subcategory_id=subcategory_id)
        except SubCategory.DoesNotExist:
            return Response({'error': 'SubCategory not found.'}, status=status.HTTP_404_NOT_FOUND)

        subcategory_data = {
            'name': request.data.get('name', subcategory.name),
            'description': request.data.get('description', subcategory.description),
            'visible': request.data.get('visible', subcategory.visible),
            'header': request.data.get('header', subcategory.header),
            'category_page': request.data.get('category_page', subcategory.category_page),
        }

        subcategory_serializer = SubCategorySerializer(subcategory, data=subcategory_data, partial=True)

        if not subcategory_serializer.is_valid():
            return Response(subcategory_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        subcategory_serializer.save()

        # Handling category updates
        category_ids = request.data.get('categories', [])
        if isinstance(category_ids, str):
            try:
                category_ids = list(map(int, category_ids.strip("[]").split(",")))
            except ValueError:
                return Response({'error': 'Invalid categories format.'}, status=status.HTTP_400_BAD_REQUEST)

        if category_ids:
            categories = Category.objects.filter(category_id__in=category_ids)
            if categories.count() != len(category_ids):
                return Response({'error': 'One or more categories not found.'}, status=status.HTTP_400_BAD_REQUEST)
            subcategory.categories.set(categories)  # Update categories

        return Response({'subcategory': subcategory_serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class SubCategoryImageUpdateView(APIView):
    def put(self, request, image_id):
        try:
            subcategory_image = SubCategoryImage.objects.get(subcategory_image_id=image_id)
        except SubCategoryImage.DoesNotExist:
            return Response({'error': 'SubCategory image not found.'}, status=status.HTTP_404_NOT_FOUND)

        if 'image' in request.FILES:
            if subcategory_image.image:
                public_id = subcategory_image.image.public_id
                result = destroy(public_id, invalidate=True)
                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete existing image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            subcategory_image.image = request.FILES['image']

        subcategory_image.alt_text = request.data.get('alt_text', subcategory_image.alt_text)
        subcategory_image.subcategory_id = request.data.get('subcategory', subcategory_image.subcategory.subcategory_id)
        subcategory_image.visible = request.data.get('visible', subcategory_image.visible)

        subcategory_image.save()
        subcategory_image_serializer = SubCategoryImageSerializer(subcategory_image)
        return Response({'subcategory_image': subcategory_image_serializer.data}, status=status.HTTP_200_OK)

@method_decorator(csrf_exempt, name='dispatch')
class SubCategoryDeleteView(APIView):
    def delete(self, request, subcategory_id):
        try:
            subcategory = SubCategory.objects.get(pk=subcategory_id)
            subcategory.delete()
            return Response({'message': 'SubCategory deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
        except SubCategory.DoesNotExist:
            return Response({'error': 'SubCategory not found.'}, status=status.HTTP_404_NOT_FOUND)

@method_decorator(csrf_exempt, name='dispatch')
class SubCategoryImageDeleteView(APIView):
    def delete(self, request, subcategory_image_id):
        try:
            subcategory_image = SubCategoryImage.objects.get(pk=subcategory_image_id)

            if subcategory_image.image:
                public_id = subcategory_image.image.public_id
                result = destroy(public_id, invalidate=True)

                if result.get('result') != 'ok':
                    return Response({'error': 'Failed to delete image from Cloudinary.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            subcategory_image.delete()
            return Response({'message': 'SubCategory image deleted successfully from both database and Cloudinary.'}, status=status.HTTP_204_NO_CONTENT)

        except SubCategoryImage.DoesNotExist:
            return Response({'error': 'SubCategory image not found.'}, status=status.HTTP_404_NOT_FOUND)

class SubCategoryDetailView(APIView):
    def get(self, request, subcategory_id):
        try:
            subcategory = SubCategory.objects.get(subcategory_id=subcategory_id)
            serializer = SubCategorySerializer(subcategory)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except SubCategory.DoesNotExist:
            return Response({'error': 'SubCategory not found.'}, status=status.HTTP_404_NOT_FOUND)

class ListSubCategoryView(APIView):
    def get(self, request):
        subcategories = SubCategory.objects.all()
        serializer = SimpleSubCategorySerializer(subcategories, many=True)
        return Response({'subcategories': serializer.data}, status=status.HTTP_200_OK)

